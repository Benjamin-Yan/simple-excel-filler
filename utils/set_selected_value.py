import os
import tempfile
from flask import session
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from shared import created_temp_files

def get_number_or_alpha(a, m):
    """ m=0 get number part, m=1 get alpha part
    a = "BAS" #x
    b = "456" #x
    c = "456ABS" #O
    d = "ABS456" #x
    e = "" #x
    """
    for i in range(len(a)):
        if (a[i]).isalpha():
            if m == 0 and i <= 0:
                return "0"
            elif m == 0:
                return a[:i]
            elif m == 1:
                return a[i:]
    return "0"

def sheet_title_checker(title):
    """ Ensure title is 31 characters or fewer """
    if len(title) > 31:
        return title[:31]
    return title

def get_column_name(infile, outfile):
    """ 說明
    回傳要讀的 excel 檔案的列數和行標題
    並同時創建暫存檔案供 fill_excel_select 函數使用
    """
    try:
        df = pd.read_excel(infile, sheet_name=0, engine='openpyxl')
        rowcnt = df.shape[0]
        columns = df.columns.tolist()

        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_fileI:
            infile.seek(0) # Reset file pointer to the beginning after reading above
            file_content = infile.read()
            if not file_content:
                raise ValueError("Uploaded file is empty.")
            
            temp_fileI.write(file_content)
            temp_fileI.close()

            print(f"Temporary fileI saved at: {temp_fileI.name}")

        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_fileO:
            outfile.seek(0)
            file_content = outfile.read()
            if not file_content:
                raise ValueError("Uploaded file is empty.")

            temp_fileO.write(file_content)
            temp_fileO.close()

            print(f"Temporary fileO saved at: {temp_fileO.name}")

        session['temp_fileI_path'] = temp_fileI.name
        session['temp_fileO_path'] = temp_fileO.name
        session['fileO_name'] = outfile.filename

        created_temp_files.append(temp_fileI.name)
        created_temp_files.append(temp_fileO.name)

        return {
            "rows": rowcnt,
            "columns": columns,
            "status": "success"
        }
    except Exception as e:
        return {
            "status": "error",
            "message": str(e)
        }

def fill_excel_select(selected_cols, selected_rows, location_data):
    """ 說明
    從第一個人填到第 selected_rows 個人
    填 selected_cols[i] 欄位的資料到對應的 location_data[i] 位置

    工作表名稱預設為 selected_cols[0] (即工讀生姓名)
    """
    # 選擇範圍 2 ~ rowcnt+1 (get and see 2~38)
    # 要去掉尾部未命名的 "Unnamed: 22" 0, 2, 5, 10, 13 (get 0~22, see A~...)
    temp_fileI_path = session['temp_fileI_path']
    session.pop('temp_fileI_path', None)
    if not temp_fileI_path or not os.path.exists(temp_fileI_path):
        return 'No file to process', 404

    temp_fileO_path = session['temp_fileO_path']
    session.pop('temp_fileO_path', None)
    if not temp_fileO_path or not os.path.exists(temp_fileO_path):
        return 'No file to process', 404

    fileO_name = session['fileO_name']

    DATA_DIR = os.path.join(os.getcwd(), "data")
    os.makedirs(DATA_DIR, exist_ok=True)
    file_path = os.path.join(DATA_DIR, fileO_name)

    try:
        df = pd.read_excel(temp_fileI_path, sheet_name=0, engine='openpyxl', dtype=str)
        rowcnt = df.shape[0]

        selected_columns_data = df.iloc[:int(selected_rows)-1, list(map(int, selected_cols))].astype(str)
        data = selected_columns_data.to_numpy().tolist() # 二維列表

        wb = load_workbook(temp_fileO_path)
        for i in data:
            ##### 填入自選資料 by each row(person) #####
            sheet_names = wb.sheetnames
            last_sheet_name = sheet_names[-1]
            last_sheet = wb[last_sheet_name]

            new_sheet = wb.copy_worksheet(last_sheet)
            new_sheet.title = sheet_title_checker(i[0]) # 工讀生名稱

            for j in range(len(i)):
                row_loc_number = get_number_or_alpha(location_data[j], 0)
                column_loc_alpha = get_number_or_alpha(location_data[j], 1)

                if row_loc_number == "0" or column_loc_alpha == "0":
                    raise ValueError("Invalid input. Please enter a valid Excel cell reference (e.g., '3F' or '27D').")
                
                column_loc_number = column_index_from_string(column_loc_alpha)

                new_sheet.cell(row=int(row_loc_number), column=int(column_loc_number), value=i[j])

        wb.save(file_path)

        return {
            "rows": rowcnt,
            "fileO_name": fileO_name,
            "status": "success"
        }
    except Exception as e:
        return {
            "status": "error",
            "message": str(e)
        }

