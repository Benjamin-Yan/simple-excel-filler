import os
import io
import pandas as pd
from openpyxl import load_workbook
from utils.set_selected_value import sheet_title_checker

def fill_excel_default(infile, outfile):
    """ 說明
    將 姓名、學號、電話、身分證 欄位填入表單中，只要是有名字的都會填入
    """
    try:
        df = pd.read_excel(infile, sheet_name=0, engine='openpyxl')
        row_count = df.shape[0]
        col_count = df.shape[1]

        # 3: 姓名, 8: 學號, 11: 電話, 12: 身分證
        selected_data = df.iloc[:, 3:13]
        selected_columns_data = selected_data.iloc[:, [0, 5, 9, 8]].astype(str)
        # or selected_columns_data = df.iloc[1:34, [3, 8, 12, 11]]

        # 將資料轉換成二維列表
        data = selected_columns_data.to_numpy().tolist()

        DATA_DIR = os.path.join(os.getcwd(), "data")
        os.makedirs(DATA_DIR, exist_ok=True)
        file_path = os.path.join(DATA_DIR, outfile.filename)

        # Use BytesIO to handle the uploaded file as a file-like object
        wb = load_workbook(io.BytesIO(outfile.read()))  # file.read() returns the file as bytes

        for i in data:
            if i[0] == "nan": # 跳過沒有名字的，預期在這裡停止填寫
                continue

            if i[3] != "nan":
                i[3] = '0' + i[3][:-2]

            sheet_names = wb.sheetnames
            last_sheet_name = sheet_names[-1]
            last_sheet = wb[last_sheet_name]

            new_sheet = wb.copy_worksheet(last_sheet)
            new_sheet.title = sheet_title_checker(i[0]) # 工讀生名稱

            # 姓名: 放在 3 行 FG 欄位
            new_sheet.cell(row=3, column=6, value=i[0])

            # 身分證: 放在 4 行 FG 欄位
            new_sheet.cell(row=4, column=6, value=i[2])

            # 學號: 放在 3 行 i 欄位
            new_sheet.cell(row=3, column=9, value=i[1])

            # 電話: 放在 4 行 i 欄位
            new_sheet.cell(row=4, column=9, value=i[3])

        wb.save(file_path)

        return {
            "rows": row_count,
            "columns": col_count,
            "status": "success"
        }
    except Exception as e:
        return {
            "status": "error",
            "message": str(e)
        }

